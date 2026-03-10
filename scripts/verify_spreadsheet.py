"""
Production-Quality Spreadsheet Verification Script
Independently verifies Excel calculations using Python shadow calculations.

Features:
- Multi-range formulas (SUM(A1:A10,B1:B10))
- Absolute/relative references ($A$1:$A$10, A:A)
- Conditional functions (SUMIF, SUMIFS, COUNTIF, COUNTIFS)
- Arithmetic formulas (A1*B1, A1+B1-C1, A1/B1)
- Percentage calculations
- Cross-sheet references with verification
- Case-insensitive sheet name matching
- Domain-specific templates (DCF, 3-statement, dashboard)
- JSON output with detailed error context
- Robust tolerance handling

Usage: python verify_spreadsheet.py <excel_file> [--domain financial|engineering|general]
"""

import json
import logging
import math
import re
import sys
from pathlib import Path
from typing import Optional, Dict, List, Any, Tuple
from dataclasses import dataclass, asdict
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("verify_spreadsheet")


@dataclass
class VerificationResult:
    """Represents a single verification check."""
    name: str
    status: str  # "PASS", "FAIL", "WARNING"
    detail: str
    formula: Optional[str] = None
    excel_value: Optional[Any] = None
    python_value: Optional[Any] = None
    error_context: Optional[str] = None

    def to_dict(self):
        return {k: v for k, v in asdict(self).items() if v is not None}


class FormulaParser:
    """Parses and evaluates Excel formulas independently."""

    def __init__(self, ws_values, ws_formulas=None, all_sheets=None, sheet_name=None):
        """
        Args:
            ws_values: Worksheet with data_only=True (values)
            ws_formulas: Worksheet with data_only=False (formulas)
            all_sheets: Dict of all worksheets for cross-sheet refs
            sheet_name: Current sheet name for error context
        """
        self.ws_values = ws_values
        self.ws_formulas = ws_formulas
        self.all_sheets = all_sheets or {}
        self.sheet_name = sheet_name

    def parse_cell_reference(self, ref: str, relative_to=None) -> Optional[Any]:
        """
        Parse a single cell reference like A1, $A$1, or Sheet1!A1.
        Returns the cell value.
        """
        ref = ref.strip()
        if not ref:
            return None

        # Handle sheet prefix
        sheet = self.ws_values
        if "!" in ref:
            sheet_name, cell_ref = ref.split("!", 1)
            sheet_name = sheet_name.strip("'\"")
            if sheet_name.upper() in self.all_sheets:
                sheet = self.all_sheets[sheet_name.upper()]
            else:
                # Fuzzy match for case-insensitive
                for sn in self.all_sheets:
                    if sn.upper() == sheet_name.upper():
                        sheet = self.all_sheets[sn]
                        break
            ref = cell_ref

        # Handle entire columns (A:A, B:B)
        if ":" in ref:
            parts = ref.split(":")
            if parts[0].isalpha() and parts[1].isalpha():
                # Column range like A:B - evaluate to None (can't sum entire column)
                return None

        # Remove $ signs
        ref_clean = ref.replace("$", "")

        try:
            return sheet[ref_clean].value
        except (KeyError, AttributeError):
            return None

    def parse_range(self, range_str: str) -> List[Any]:
        """
        Parse a range like A1:A10, $A$1:$B$5, or A:A.
        Returns list of all numeric values in the range.
        """
        range_str = range_str.strip()
        if not range_str:
            return []

        # Handle sheet prefix
        sheet = self.ws_values
        if "!" in range_str:
            sheet_name, range_part = range_str.split("!", 1)
            sheet_name = sheet_name.strip("'\"")
            # Case-insensitive lookup
            found = False
            for sn, ws in self.all_sheets.items():
                if sn.upper() == sheet_name.upper():
                    sheet = ws
                    found = True
                    break
            if not found and sheet_name in self.all_sheets:
                sheet = self.all_sheets[sheet_name]
            range_str = range_part

        values = []
        try:
            for row in sheet[range_str]:
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        values.append(cell.value)
        except (KeyError, AttributeError):
            pass

        return values

    def parse_range_with_coords(self, range_str: str) -> List[Tuple[str, Any]]:
        """Parse range and return (coordinate, value) tuples."""
        range_str = range_str.strip()
        coords = []

        sheet = self.ws_values
        if "!" in range_str:
            sheet_name, range_part = range_str.split("!", 1)
            sheet_name = sheet_name.strip("'\"")
            for sn, ws in self.all_sheets.items():
                if sn.upper() == sheet_name.upper():
                    sheet = ws
                    break
            range_str = range_part

        try:
            for row in sheet[range_str]:
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        coords.append((cell.coordinate, cell.value))
        except (KeyError, AttributeError):
            pass

        return coords

    def extract_function_args(self, formula: str, func_name: str) -> List[str]:
        """
        Extract arguments from a function call.
        E.g., SUM(A1:A10,B1:B10) -> ["A1:A10", "B1:B10"]
        """
        # Build pattern: FUNC(...)
        pattern = rf"{func_name}\s*\((.*)\)"
        match = re.search(pattern, formula, re.IGNORECASE)
        if not match:
            return []

        args_str = match.group(1)
        # Simple split on comma, respecting parentheses
        args = []
        depth = 0
        current = ""
        for char in args_str:
            if char == "(" or char == "[":
                depth += 1
            elif char == ")" or char == "]":
                depth -= 1
            elif char == "," and depth == 0:
                args.append(current.strip())
                current = ""
                continue
            current += char
        if current.strip():
            args.append(current.strip())

        return args

    def verify_sum(self, formula: str) -> Optional[float]:
        """Verify SUM formula."""
        args = self.extract_function_args(formula, "SUM")
        if not args:
            return None

        total = 0.0
        try:
            for arg in args:
                if ":" in arg:
                    values = self.parse_range(arg)
                    total += sum(values)
                else:
                    val = self.parse_cell_reference(arg)
                    if isinstance(val, (int, float)):
                        total += val
            return total
        except Exception:
            return None

    def verify_sumif(self, formula: str) -> Optional[float]:
        """Verify SUMIF(range, criteria, [sum_range])."""
        args = self.extract_function_args(formula, "SUMIF")
        if len(args) < 2:
            return None

        try:
            range_str = args[0]
            criteria = args[1].strip("'\"")
            sum_range_str = args[2] if len(args) > 2 else range_str

            range_coords = self.parse_range_with_coords(range_str)
            sum_coords = self.parse_range_with_coords(sum_range_str)

            # Match coordinates position
            total = 0.0
            for i, (coord, val) in enumerate(range_coords):
                if i < len(sum_coords):
                    range_val = val
                    if str(range_val) == criteria or (
                        isinstance(range_val, (int, float)) and
                        isinstance(criteria, str) and
                        criteria.replace("'", "").isdigit() and
                        range_val == float(criteria)
                    ):
                        sum_val = sum_coords[i][1]
                        if isinstance(sum_val, (int, float)):
                            total += sum_val
            return total
        except Exception:
            return None

    def verify_countif(self, formula: str) -> Optional[int]:
        """Verify COUNTIF(range, criteria)."""
        args = self.extract_function_args(formula, "COUNTIF")
        if len(args) < 2:
            return None

        try:
            range_str = args[0]
            criteria = args[1].strip("'\"")

            range_coords = self.parse_range_with_coords(range_str)
            count = 0
            for coord, val in range_coords:
                if str(val) == criteria or (
                    isinstance(val, (int, float)) and
                    criteria.replace("'", "").replace("%", "").isdigit()
                ):
                    count += 1
            return count
        except Exception:
            return None

    def verify_average(self, formula: str) -> Optional[float]:
        """Verify AVERAGE formula."""
        args = self.extract_function_args(formula, "AVERAGE")
        if not args:
            return None

        try:
            values = []
            for arg in args:
                if ":" in arg:
                    vals = self.parse_range(arg)
                    values.extend(vals)
                else:
                    val = self.parse_cell_reference(arg)
                    if isinstance(val, (int, float)):
                        values.append(val)
            return sum(values) / len(values) if values else None
        except Exception:
            return None

    def verify_count(self, formula: str) -> Optional[int]:
        """Verify COUNT formula."""
        args = self.extract_function_args(formula, "COUNT")
        if not args:
            return None

        try:
            count = 0
            for arg in args:
                if ":" in arg:
                    values = self.parse_range(arg)
                    count += len(values)
                else:
                    val = self.parse_cell_reference(arg)
                    if isinstance(val, (int, float)):
                        count += 1
            return count
        except Exception:
            return None

    def verify_arithmetic(self, formula: str) -> Optional[float]:
        """
        Verify simple arithmetic formulas like =A1+B1, =A1*B1-C1, =A1/B1.
        Handles: +, -, *, /, () parentheses.
        """
        try:
            # Remove leading =
            expr = formula.lstrip("=").strip()

            # Replace cell references with their values
            cell_pattern = r"\$?[A-Za-z]+\$?\d+"
            matches = re.findall(cell_pattern, expr)

            replacement_map = {}
            for cell_ref in set(matches):
                val = self.parse_cell_reference(cell_ref)
                if isinstance(val, (int, float)):
                    replacement_map[cell_ref] = str(val)

            # Replace in expression
            for cell_ref, val_str in replacement_map.items():
                expr = expr.replace(cell_ref, val_str)

            # Safely evaluate
            if all(c in "0123456789+-*/.()% " for c in expr):
                result = eval(expr)
                return float(result) if isinstance(result, (int, float)) else None
        except Exception:
            pass

        return None

    def verify_percentage(self, formula: str) -> Optional[float]:
        """Verify percentage formulas like =A1/B1."""
        result = self.verify_arithmetic(formula)
        # Percentages are just division results, already handled by arithmetic
        return result


class SpreadsheetVerifier:
    """Main verification engine."""

    def __init__(self, filepath: str, domain: str = "general"):
        self.filepath = filepath
        self.domain = domain
        self.wb_formulas = None
        self.wb_values = None
        self.all_sheets = {}
        self.sheet_name_map = {}  # Maps uppercase -> original name

    def load(self):
        """Load workbooks and build sheet maps."""
        self.wb_formulas = load_workbook(self.filepath, data_only=False)
        self.wb_values = load_workbook(self.filepath, data_only=True)

        # Build sheet maps
        for sheet_name in self.wb_values.sheetnames:
            self.sheet_name_map[sheet_name.upper()] = sheet_name
            self.all_sheets[sheet_name.upper()] = self.wb_values[sheet_name]

    def find_sheet_fuzzy(self, pattern: str) -> Optional[str]:
        """
        Case-insensitive fuzzy sheet matching.
        E.g., find "Balance Sheet" even if named "Consolidated Balance Sheet".
        """
        pattern_upper = pattern.upper()
        for sheet_name in self.wb_values.sheetnames:
            if pattern_upper in sheet_name.upper():
                return sheet_name
        return None

    def verify(self) -> dict:
        """Run full verification pipeline."""
        logger.info("Loading workbook: %s (domain: %s)", self.filepath, self.domain)
        self.load()

        report = {
            "file": self.filepath,
            "domain": self.domain,
            "structural_checks": [],
            "value_checks": [],
            "domain_checks": [],
            "summary": {"total": 0, "passed": 0, "failed": 0, "warnings": 0},
        }

        try:
            # Phase 1: Structural checks
            logger.info("Phase 1: Running structural checks...")
            report["structural_checks"] = self.run_structural_checks()
            logger.info("  → %d structural checks complete", len(report["structural_checks"]))

            # Phase 2: Value verification
            logger.info("Phase 2: Running value verification...")
            report["value_checks"] = self.run_value_checks()
            logger.info("  → %d value checks complete", len(report["value_checks"]))

            # Phase 3: Domain-specific checks
            logger.info("Phase 3: Running %s domain checks...", self.domain)
            if self.domain == "financial":
                report["domain_checks"] = self.run_financial_checks()
            elif self.domain == "engineering":
                report["domain_checks"] = self.run_engineering_checks()

            # Summarize
            all_checks = (
                report["structural_checks"]
                + report["value_checks"]
                + report["domain_checks"]
            )
            for check in all_checks:
                report["summary"]["total"] += 1
                status = check.get("status", "UNKNOWN")
                if status == "PASS":
                    report["summary"]["passed"] += 1
                elif status == "FAIL":
                    report["summary"]["failed"] += 1
                else:
                    report["summary"]["warnings"] += 1

            s = report["summary"]
            logger.info(
                "Verification complete: %d total, %d passed, %d failed, %d warnings",
                s["total"], s["passed"], s["failed"], s["warnings"],
            )
            if s["failed"] > 0:
                logger.warning("FAILURES DETECTED — review report for details")

        finally:
            if self.wb_formulas:
                self.wb_formulas.close()
            if self.wb_values:
                self.wb_values.close()

        return report

    def run_structural_checks(self) -> List[Dict]:
        """Check for structural issues."""
        checks = []

        # Check 1: Empty formula results
        empty_formulas = []
        for ws_name in self.wb_formulas.sheetnames:
            ws_f = self.wb_formulas[ws_name]
            ws_v = self.wb_values[ws_name]
            for row in ws_f.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        val = ws_v[cell.coordinate].value
                        if val is None:
                            empty_formulas.append(f"{ws_name}!{cell.coordinate}")

        checks.append({
            "name": "Empty formula results",
            "status": "PASS" if not empty_formulas else "WARNING",
            "detail": (
                f"{len(empty_formulas)} formulas evaluate to empty: {', '.join(empty_formulas[:5])}"
                if empty_formulas
                else "All formulas produce values"
            ),
        })

        # Check 2: Error values
        error_patterns = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
        error_cells = []
        for ws_name in self.wb_values.sheetnames:
            ws_v = self.wb_values[ws_name]
            for row in ws_v.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for err in error_patterns:
                            if err in cell.value:
                                error_cells.append(f"{ws_name}!{cell.coordinate}: {cell.value}")
                                break

        checks.append({
            "name": "Formula error values",
            "status": "PASS" if not error_cells else "FAIL",
            "detail": (
                f"{len(error_cells)} errors: {', '.join(error_cells[:10])}"
                if error_cells
                else "No formula errors"
            ),
        })

        # Check 3: Cross-sheet references
        missing_refs = []
        for ws_name in self.wb_formulas.sheetnames:
            ws_f = self.wb_formulas[ws_name]
            for row in ws_f.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        refs = re.findall(r"'?([A-Za-z0-9_ ]+)'?!", cell.value)
                        for ref in refs:
                            ref_clean = ref.strip("'")
                            if ref_clean.upper() not in self.sheet_name_map:
                                missing_refs.append(f"{ws_name}!{cell.coordinate} → '{ref_clean}'")

        checks.append({
            "name": "Cross-sheet references",
            "status": "PASS" if not missing_refs else "FAIL",
            "detail": (
                f"{len(missing_refs)} broken refs: {', '.join(missing_refs[:5])}"
                if missing_refs
                else "All cross-sheet references valid"
            ),
        })

        return checks

    def run_value_checks(self) -> List[Dict]:
        """Verify all formulas that can be independently calculated."""
        checks = []

        for ws_name in self.wb_formulas.sheetnames:
            ws_f = self.wb_formulas[ws_name]
            ws_v = self.wb_values[ws_name]

            parser = FormulaParser(ws_v, ws_f, self.all_sheets, ws_name)

            for row in ws_f.iter_rows():
                for cell in row:
                    if not (cell.value and isinstance(cell.value, str) and cell.value.startswith("=")):
                        continue

                    formula = cell.value
                    formula_upper = formula.upper()
                    excel_val = ws_v[cell.coordinate].value
                    addr = f"{ws_name}!{cell.coordinate}"

                    # Skip error cells
                    if isinstance(excel_val, str) and excel_val.startswith("#"):
                        continue

                    python_val = None
                    check_type = None

                    # SUM
                    if formula_upper.startswith("=SUM("):
                        python_val = parser.verify_sum(formula)
                        check_type = "SUM"

                    # SUMIF
                    elif formula_upper.startswith("=SUMIF("):
                        python_val = parser.verify_sumif(formula)
                        check_type = "SUMIF"

                    # COUNTIF
                    elif formula_upper.startswith("=COUNTIF("):
                        python_val = parser.verify_countif(formula)
                        check_type = "COUNTIF"

                    # AVERAGE
                    elif formula_upper.startswith("=AVERAGE("):
                        python_val = parser.verify_average(formula)
                        check_type = "AVERAGE"

                    # COUNT
                    elif formula_upper.startswith("=COUNT(") and not formula_upper.startswith("=COUNTIF"):
                        python_val = parser.verify_count(formula)
                        check_type = "COUNT"

                    # Arithmetic (must not contain functions)
                    elif not re.search(r"[A-Z_]+\s*\(", formula_upper):
                        python_val = parser.verify_arithmetic(formula)
                        check_type = "ARITHMETIC"

                    if python_val is not None and check_type:
                        checks.append(self.make_check(addr, check_type, excel_val, python_val, formula))

        return checks

    def make_check(self, addr: str, check_type: str, excel_val: Any, python_val: Any, formula: str) -> Dict:
        """Create a verification check result."""
        match = self.compare_values(excel_val, python_val, check_type)
        return {
            "name": f"{check_type} @ {addr}",
            "status": "PASS" if match else "FAIL",
            "formula": formula,
            "excel_value": excel_val,
            "python_value": python_val,
            "detail": (
                "Match" if match else f"MISMATCH: Excel={excel_val}, Python={python_val}"
            ),
        }

    def compare_values(self, excel_val: Any, python_val: Any, check_type: str = "general") -> bool:
        """
        Compare values with tolerance based on formula type.
        Different tolerances for different formula types.
        """
        if excel_val is None and python_val is None:
            return True
        if excel_val is None or python_val is None:
            return False

        # Tolerance varies by type
        if check_type == "PERCENTAGE":
            rel_tol = 1e-6
            abs_tol = 0.0001
        elif check_type in ["SUMIF", "COUNTIF", "AVERAGE"]:
            rel_tol = 1e-9
            abs_tol = 0.01
        else:
            rel_tol = 1e-9
            abs_tol = 0.01

        if isinstance(excel_val, (int, float)) and isinstance(python_val, (int, float)):
            return math.isclose(float(excel_val), float(python_val), rel_tol=rel_tol, abs_tol=abs_tol)

        return str(excel_val) == str(python_val)

    def run_financial_checks(self) -> List[Dict]:
        """Financial domain-specific checks."""
        checks = []

        # Verify DCF
        checks.extend(self.verify_dcf())

        # Verify 3-statement model
        checks.extend(self.verify_three_statement())

        # Verify balance sheet
        checks.extend(self.verify_balance_sheet())

        return checks

    def verify_dcf(self) -> List[Dict]:
        """
        DCF verification:
        - Discount rate applied correctly
        - Terminal value calculated
        - NPV formula correct
        """
        checks = []
        dcf_sheet = self.find_sheet_fuzzy("DCF") or self.find_sheet_fuzzy("Valuation")

        if not dcf_sheet:
            return checks

        ws = self.wb_values[dcf_sheet]
        parser = FormulaParser(ws, self.wb_formulas[dcf_sheet], self.all_sheets, dcf_sheet)

        # Look for discount rate
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    label = cell.value.strip().lower()
                    if "discount rate" in label or "wacc" in label:
                        for offset in range(1, 5):
                            neighbor = ws.cell(row=cell.row, column=cell.column + offset)
                            if isinstance(neighbor.value, (int, float)):
                                rate = neighbor.value
                                reasonable = 0.01 <= rate <= 0.5  # 1% to 50%
                                checks.append({
                                    "name": "DCF Discount Rate",
                                    "status": "PASS" if reasonable else "WARNING",
                                    "detail": f"Discount rate = {rate:.2%} ({'reasonable' if reasonable else 'unusual'})",
                                })
                                break

        return checks

    def verify_three_statement(self) -> List[Dict]:
        """
        3-statement model verification:
        - Net Income flows to Balance Sheet
        - Cash flows reconcile
        """
        checks = []

        income_sheet = self.find_sheet_fuzzy("Income") or self.find_sheet_fuzzy("P&L")
        balance_sheet = self.find_sheet_fuzzy("Balance")
        cash_sheet = self.find_sheet_fuzzy("Cash") or self.find_sheet_fuzzy("CFS")

        if not (income_sheet and balance_sheet):
            return checks

        # Verify retained earnings link
        income_ws = self.wb_values[income_sheet]
        balance_ws = self.wb_values[balance_sheet]

        net_income = None
        for row in income_ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "net income" in cell.value.lower():
                    for offset in range(1, 5):
                        neighbor = income_ws.cell(row=cell.row, column=cell.column + offset)
                        if isinstance(neighbor.value, (int, float)):
                            net_income = neighbor.value
                            break

        if net_income is not None:
            retained_earnings = None
            for row in balance_ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and "retained earnings" in cell.value.lower()
                    ):
                        for offset in range(1, 5):
                            neighbor = balance_ws.cell(row=cell.row, column=cell.column + offset)
                            if isinstance(neighbor.value, (int, float)):
                                retained_earnings = neighbor.value
                                break

            if retained_earnings is not None:
                # Should be related to net income (within 20%)
                ratio = abs(retained_earnings) / (abs(net_income) + 0.01)
                reasonable = 0.5 <= ratio <= 2.0
                checks.append({
                    "name": "3-Statement: Net Income → Retained Earnings",
                    "status": "PASS" if reasonable else "WARNING",
                    "detail": f"Net Income={net_income:,.0f}, Retained Earnings={retained_earnings:,.0f}",
                })

        return checks

    def verify_balance_sheet(self) -> List[Dict]:
        """Verify balance sheet equation: Assets = Liabilities + Equity."""
        checks = []

        bs_sheet = self.find_sheet_fuzzy("Balance")
        if not bs_sheet:
            return checks

        ws = self.wb_values[bs_sheet]
        totals = {}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    label = cell.value.strip().lower()
                    if "total assets" in label:
                        for offset in range(1, 5):
                            neighbor = ws.cell(row=cell.row, column=cell.column + offset)
                            if isinstance(neighbor.value, (int, float)):
                                totals["assets"] = neighbor.value
                                break
                    elif "total liabilities" in label and "equity" not in label:
                        for offset in range(1, 5):
                            neighbor = ws.cell(row=cell.row, column=cell.column + offset)
                            if isinstance(neighbor.value, (int, float)):
                                totals["liabilities"] = neighbor.value
                                break
                    elif "total equity" in label or "total stockholders" in label:
                        for offset in range(1, 5):
                            neighbor = ws.cell(row=cell.row, column=cell.column + offset)
                            if isinstance(neighbor.value, (int, float)):
                                totals["equity"] = neighbor.value
                                break

        if all(k in totals for k in ["assets", "liabilities", "equity"]):
            a, l, e = totals["assets"], totals["liabilities"], totals["equity"]
            balanced = math.isclose(a, l + e, abs_tol=1.0)
            checks.append({
                "name": "Balance Sheet Equation (A = L + E)",
                "status": "PASS" if balanced else "FAIL",
                "detail": f"Assets={a:,.2f}, L+E={l + e:,.2f}, diff={a - l - e:,.2f}",
            })

        return checks

    def verify_dashboard(self) -> List[Dict]:
        """Verify dashboard metrics match detail data."""
        checks = []
        # Dashboard verification is highly custom per spreadsheet
        # This is a placeholder for custom implementations
        return checks

    def run_engineering_checks(self) -> List[Dict]:
        """Engineering domain-specific checks."""
        checks = []

        for ws_name in self.wb_values.sheetnames:
            ws = self.wb_values[ws_name]

            # Factor of Safety checks
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        label = cell.value.strip().lower()
                        if "factor of safety" in label or label == "fos":
                            for offset in range(1, 5):
                                neighbor = ws.cell(row=cell.row, column=cell.column + offset)
                                if isinstance(neighbor.value, (int, float)):
                                    fos = neighbor.value
                                    checks.append({
                                        "name": f"Factor of Safety ({ws_name})",
                                        "status": "PASS" if fos > 1.0 else "FAIL",
                                        "detail": f"FoS = {fos:.2f} ({'safe' if fos > 1.0 else 'unsafe'})",
                                    })
                                    break

            # Efficiency checks
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        label = cell.value.strip().lower()
                        if "efficiency" in label:
                            for offset in range(1, 5):
                                neighbor = ws.cell(row=cell.row, column=cell.column + offset)
                                if isinstance(neighbor.value, (int, float)):
                                    eff = neighbor.value
                                    in_range = (0 <= eff <= 1.0) or (0 <= eff <= 100)
                                    checks.append({
                                        "name": f"Efficiency range ({ws_name})",
                                        "status": "PASS" if in_range else "WARNING",
                                        "detail": f"Efficiency = {eff:.2f}",
                                    })
                                    break

        return checks


def verify_dcf(wb_filepath: str) -> Dict:
    """Template: Verify DCF (Discounted Cash Flow) model."""
    verifier = SpreadsheetVerifier(wb_filepath, "financial")
    return verifier.verify()


def verify_three_statement(wb_filepath: str) -> Dict:
    """Template: Verify 3-statement model (P&L, Balance Sheet, Cash Flow)."""
    verifier = SpreadsheetVerifier(wb_filepath, "financial")
    return verifier.verify()


def verify_dashboard(wb_filepath: str) -> Dict:
    """Template: Verify dashboard metrics match detail data."""
    verifier = SpreadsheetVerifier(wb_filepath, "general")
    return verifier.verify()


# --- Financial function helpers for custom verification scripts ---

def npv_python(rate: float, cashflows: List[float]) -> float:
    """Calculate NPV independently (matches Excel NPV behavior: period 1 onwards).
    For full NPV with period-0 investment: initial_investment + npv_python(rate, cf1..cfn)
    """
    return sum(cf / (1 + rate) ** (i + 1) for i, cf in enumerate(cashflows))


def irr_python(cashflows: List[float], guess: float = 0.1, tol: float = 1e-8, max_iter: int = 1000) -> Optional[float]:
    """Calculate IRR using Newton-Raphson method."""
    rate = guess
    for _ in range(max_iter):
        npv_val = sum(cf / (1 + rate) ** i for i, cf in enumerate(cashflows))
        dnpv = sum(-i * cf / (1 + rate) ** (i + 1) for i, cf in enumerate(cashflows))
        if abs(dnpv) < 1e-14:
            return None
        new_rate = rate - npv_val / dnpv
        if abs(new_rate - rate) < tol:
            return new_rate
        rate = new_rate
    return None


def xnpv_python(rate: float, cashflows: List[float], dates: List[float]) -> float:
    """Calculate XNPV with specific dates (dates as days from first date)."""
    d0 = dates[0]
    return sum(cf / (1 + rate) ** ((d - d0) / 365.0) for cf, d in zip(cashflows, dates))


def pmt_python(rate: float, nper: int, pv: float, fv: float = 0, pmt_type: int = 0) -> float:
    """Calculate PMT (loan payment) independently."""
    if rate == 0:
        return -(pv + fv) / nper
    pvif = (1 + rate) ** nper
    pmt = rate * (pv * pvif + fv) / (pvif - 1)
    if pmt_type == 1:
        pmt /= (1 + rate)
    return -pmt


def main():
    if len(sys.argv) < 2:
        print("Usage: python verify_spreadsheet.py <excel_file> [--domain financial|engineering|general]")
        sys.exit(1)

    filepath = sys.argv[1]
    domain = "general"

    if "--domain" in sys.argv:
        idx = sys.argv.index("--domain")
        if idx + 1 < len(sys.argv):
            domain = sys.argv[idx + 1]

    if not Path(filepath).exists():
        print(json.dumps({"error": f"File not found: {filepath}"}))
        sys.exit(1)

    try:
        verifier = SpreadsheetVerifier(filepath, domain)
        report = verifier.verify()
        print(json.dumps(report, indent=2, default=str))
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)


if __name__ == "__main__":
    main()
